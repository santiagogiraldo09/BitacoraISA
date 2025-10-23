from flask import Flask, request, jsonify, render_template, send_file, redirect,url_for, flash, jsonify
import azure.cognitiveservices.speech as speechsdk
from azure.storage.blob import BlobServiceClient,BlobClient,ContainerClient
from werkzeug.utils import secure_filename
import base64
import io
from io import BytesIO
from PIL import Image
import os
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from flask_cors import CORS
from datetime import datetime
from azure.storage.blob import ContentSettings
from dotenv import load_dotenv
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.lists.list import List
from office365.sharepoint.listitems.listitem import ListItem
import psycopg2
from werkzeug.security import generate_password_hash, check_password_hash
from flask import session
import secrets
from pydub import AudioSegment
import tempfile
import traceback
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
import base64

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Configuración PostgreSQL
POSTGRES_CONFIG = {
    "host": "localhost",
    "database": "Bitacora",
    "user": "postgres",  # Normalmente 'postgres' por defecto
    "password": "Daniel2030#",
    "port": "5432"  # Puerto predeterminado de PostgreSQL
}

# Configura SharePoint (modifica con tus datos)
SHAREPOINT_SITE_URL = "https://iacsas.sharepoint.com/sites/Pruebasproyectossantiago"
LIST_NAME = "Proyectos"  # Nombre de la biblioteca
LIST_NAME_REGISTROS = "RegistrosBitacora"
SHAREPOINT_USER = "santiago.giraldo@iac.com.co"
SHAREPOINT_PASSWORD = "Latumbanuncamuere3"


# Cargar variables de entorno
#load_dotenv('config/settings.env')  # Ruta relativa al archivo .env

app = Flask(__name__,template_folder='templates')
app.secret_key = secrets.token_hex(16)  # Clave secreta para sesiones
#app.secret_key = '78787878tyg8987652vgdfdf3445'
CORS(app)

projects = []

# Conecta con el servicio de Blob Storage de Azure
connection_string = "DefaultEndpointsProtocol=https;AccountName=registrobitacora;AccountKey=ZyHZAOvOBijiOfY3BR3ZEDZsCAHOu3swEPnS+D7AacR2Yr94HS+jBMa2/20sJpZ71decGXYHQxE2+AStBWI/wA==;EndpointSuffix=core.windows.net"
container_name = "registros"


# Inicializa el cliente de BlobServiceClient
blob_service_client = BlobServiceClient.from_connection_string(connection_string)

def create_user(nombre, apellido, email, password, cargo, rol, empresa):
    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()
        
        hashed_password = generate_password_hash(password)
        
        cursor.execute(
            """INSERT INTO usuario (name, apellido, email, password, cargo, rol, empresa)
               VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING user_id""",
            (nombre, apellido, email, hashed_password, cargo, rol, empresa)
        )
        
        user_id = cursor.fetchone()[0]
        conn.commit()
        return user_id
    except psycopg2.Error as e:
        print(f"Error al crear usuario: {e}")
        return None
    finally:
        if conn:
            conn.close()

def verify_user(email, password):
    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()
        
        cursor.execute(
            "SELECT user_id, password FROM usuario WHERE email = %s",
            (email,)
        )
        
        user = cursor.fetchone()
        if user and check_password_hash(user[1], password):
            return user[0]  # Devuelve el ID del usuario
        return None
    except psycopg2.Error as e:
        print(f"Error al verificar usuario: {e}")
        return None
    finally:
        if conn:
            conn.close()


def insert_registro_bitacora(respuestas, id_proyecto, fotos=None, videos=None):
    """
    Inserta un nuevo registro de bitácora, junto con sus fotos y videos asociados
    y sus descripciones, en la base de datos.
    """
    conn = None  # Definimos conn aquí para asegurarnos de que exista en el bloque finally
    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()

        # CAMBIO 1: Simplificamos el INSERT principal.
        # - Eliminamos la columna 'foto_base64' que ya es obsoleta.
        # - Cambiamos los nombres de las claves para que coincidan con tu formulario.
        cursor.execute("""
            INSERT INTO registrosbitacoraeqing (
                zona_intervencion, -- Mapeado desde "Tipo de informe"
                items,             -- Mapeado desde "Sede"
                metros_lineales,   -- Mapeado desde "Repuestos utilizados"
                proximas_tareas,   -- Mapeado desde "Repuestos a cotizar"
                id_proyecto
            )
            VALUES (%s, %s, %s, %s, %s)
            RETURNING id_registro
        """, (
            respuestas.get('zona_intervencion'),
            respuestas.get('items'),
            respuestas.get('metros_lineales'),
            respuestas.get('proximas_tareas'),
            id_proyecto,
        ))
        id_registro = cursor.fetchone()[0]

        # CAMBIO 2: Actualizamos el bucle para que maneje objetos (archivo + descripción).
        # Ahora esperamos una lista de diccionarios, no solo una lista de strings.
        for foto_obj in fotos or []:
            file_data = foto_obj.get('file_data')
            description = foto_obj.get('description')
            cursor.execute(
                """INSERT INTO fotos_registro 
                   (id_registro, imagen_base64, description) 
                   VALUES (%s, %s, %s)""",
                (id_registro, file_data, description)
            )

        # CAMBIO 3: Hacemos lo mismo para los videos.
        for video_obj in videos or []:
            file_data = video_obj.get('file_data')
            description = video_obj.get('description')
            cursor.execute(
                """INSERT INTO videos_registro 
                   (id_registro, video_base64, description) 
                   VALUES (%s, %s, %s)""",
                (id_registro, file_data, description)
            )

        conn.commit()
        print(f"Registro {id_registro} guardado exitosamente en PostgreSQL.")

    except psycopg2.Error as e: # MEJORA: Capturamos el error específico de psycopg2 para más detalles
        print(f"Error de base de datos al guardar en PostgreSQL: {e}")
        # Opcional: podrías querer que la función devuelva un error
        # raise e 
    except Exception as e:
        print(f"Error general al guardar en PostgreSQL: {str(e)}")
        # raise e
    finally:
        if conn:
            conn.close()

def create_project(user_id, nombre, fecha_inicio, fecha_fin, director, ubicacion, coordenadas):
    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        #conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute(
            """INSERT INTO proyectos (nombre_proyecto, fecha_inicio, fecha_fin, director_obra, ubicacion, coordenadas, user_id)
               VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id_proyecto""",
            (nombre, fecha_inicio, fecha_fin, director, ubicacion, coordenadas, user_id)
        )
        
        project_id = cursor.fetchone()[0]
        conn.commit()
        return project_id
    except psycopg2.Error as e:
        print(f"Error al crear proyecto: {e}")
        return None
    finally:
        if conn:
            conn.close()

def get_user_projects(user_id):
    conn = None
    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        #conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute(
            """SELECT id_proyecto, nombre_proyecto, fecha_inicio, director_obra, user_id 
               FROM proyectos WHERE user_id = %s ORDER BY fecha_inicio DESC""",
            (user_id,)
        )
        
        projects = []
        for row in cursor.fetchall():
            projects.append({
                'id_proyecto': row[0],
                'name': row[1],
                'fecha_inicio': row[2].strftime('%Y-%m-%d'),
                'director_obra': row[3],
                'user_id': row[4],

            })
        
        return projects
    except psycopg2.Error as e:
        print(f"Error al obtener proyectos: {e}")
        return []
    finally:
        if conn:
            conn.close()

# Función para subir archivos a Azure Blob Storage
def upload_to_blob(file_name, data, content_type):
    try:
        blob_client = blob_service_client.get_blob_client(container=container_name, blob=file_name)
        blob_client.upload_blob(data, blob_type="BlockBlob", content_settings={"content_type": content_type})
        print(f"Archivo {file_name} subido con éxito.")
    except Exception as e:
        print(f"Error al subir {file_name}: {e}")
        raise


def get_speech_config():
    speech_key = '999fcb4d3f34436ab454ec47920febe0'
    service_region = 'centralus'
    speech_config = speechsdk.SpeechConfig(subscription=speech_key, region=service_region)
    speech_config.speech_recognition_language = "es-CO"
    speech_config.speech_synthesis_language = "es-CO"
    speech_config.speech_synthesis_voice_name = "es-CO-GonzaloNeural"
    speech_config.set_property(speechsdk.PropertyId.SpeechServiceConnection_EndSilenceTimeoutMs, "8000")
    return speech_config

def synthesize_speech(text):
    speech_config = get_speech_config()
    audio_config = speechsdk.audio.AudioOutputConfig(use_default_speaker=True)
    synthesizer = speechsdk.SpeechSynthesizer(speech_config=speech_config, audio_config=audio_config)
    result = synthesizer.speak_text_async(text).get()
    return result.reason == speechsdk.ResultReason.SynthesizingAudioCompleted

#Obtener los proyectos desde Azure Blob Storage
def get_projects_from_blob():
    projects = []
    try:
        # Obtener el cliente del contenedor
        container_client = blob_service_client.get_container_client(container_name)
        
        # Listar los blobs en el directorio de proyectos
        blobs = list(container_client.list_blobs(name_starts_with="Proyectos/"))
        
        for blob in blobs:
            if blob.name.endswith('.txt'):
                # Obtener el cliente del blob
                blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob.name)
                
                # Descargar el contenido del blob
                content = blob_client.download_blob().readall().decode('utf-8')
                
                # Extraer información del proyecto
                project_info = {}
                for line in content.strip().split('\n'):
                    line = line.strip()
                    if line:
                        parts = line.split(':', 1)
                        if len(parts) == 2:
                            key = parts[0].strip()
                            value = parts[1].strip()
                            project_info[key] = value
                
                # Extraer el nombre del proyecto del nombre del archivo
                file_name = blob.name.split('/')[-1]
                project_name = file_name.replace('proyecto_', '').replace('.txt', '')
                
                # Crear un objeto de proyecto
                project = {
                    'name': project_info.get('Nombre del Proyecto', project_name),
                    'date': project_info.get('Fecha de Inicio', 'Fecha no disponible'),
                    'blob_name': blob.name,
                    # Añadir más campos según sea necesario
                }
                
                projects.append(project)
                
    except Exception as e:
        print(f"Error al obtener proyectos del Blob Storage: {e}")
    
    return projects

@app.after_request
def add_header(response):
    response.headers["ngrok-skip-browser-warning"] = "true"
    return response

@app.route('/')
def principalscreen():
    return render_template('PrincipalScreen.html')

@app.route('/paginaprincipal')
def paginaprincipal():
    if 'user_id' not in session:
        return redirect(url_for('principalscreen'))
    
    project_id = request.args.get('project_id')
    if project_id:
        # Verificar que el proyecto pertenece al usuario
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT 1 FROM proyectos WHERE id_proyecto = %s AND user_id = %s",
            (project_id, session['user_id'])
        )
        if not cursor.fetchone():
            flash('No tienes acceso a este proyecto', 'error')
            return redirect(url_for('history'))
        conn.close()
    
    return render_template('paginaprincipal.html')

@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        apellido = request.form.get('apellido')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        empresa = request.form.get('empresa')
        cargo = request.form.get('cargo')
        rol = request.form.get('rol')
        
        if password != confirm_password:
            flash('Las contraseñas no coinciden', 'error')
            return redirect(url_for('registro'))
        
        user_id = create_user(nombre, apellido, email, password, empresa, cargo, rol)
        if user_id:
            flash('Registro exitoso. Por favor inicie sesión.', 'success')
            return redirect(url_for('principalscreen'))
        else:
            flash('Error al registrar el usuario', 'error')
    
    return render_template('registro.html')

@app.route('/login', methods=['POST'])
def login():
    email = request.form.get('email')
    password = request.form.get('password')
    
    # Validación básica de campos vacíos
    if not email or not password:
        flash('Por favor ingrese ambos campos: email y contraseña', 'error')
        return redirect(url_for('principalscreen'))

    user_id = verify_user(email, password)
    if user_id:
        # Aquí puedes implementar sesiones o JWT
        session['user_id'] = user_id #Establecer sesión
        flash('Inicio de sesión exitoso', 'success')
        return redirect(url_for('registros'))
    else:
        flash('Email o contraseña incorrectos', 'error')
        return redirect(url_for('principalscreen'))

@app.route('/index')
def index():
    return render_template('index.html')

@app.route('/registros')
def registros():
    if 'user_id' not in session:
        return redirect(url_for('principalscreen'))
    
    # Obtener proyectos de PostgreSQL
    db_projects = get_user_projects(session['user_id'])
    
    # Obtener proyectos de Azure Blob (si aún los necesitas)
    #blob_projects = get_projects_from_blob()  # Tu función existente
    
    # Combinar proyectos (o usar solo los de PostgreSQL)
    return render_template('registros.html', 
                         db_projects=db_projects)

# Ruta para la vista "history"
@app.route('/history')
def history():
    # Obtener proyectos del Blob Storage
    #blob_projects = get_projects_from_blob()
    # Obtener proyectos de PostgreSQL
    db_projects = get_user_projects(session['user_id'])
    
    # Obtener proyectos de Azure Blob (si aún los necesitas)
    #blob_projects = get_projects_from_blob()  # Tu función existente
    
    # Combinar proyectos (o usar solo los de PostgreSQL)
    return render_template('history.html', 
                         db_projects=db_projects)

@app.route('/usuario')
def usuario():
    return render_template('usuario.html')

@app.route('/inventario')
def inventario():
    return render_template('inventario.html')

# En tu archivo app.py

@app.route('/historialRegistro')
def historialregistro():
    project_id = request.args.get('project_id')
    project_name = request.args.get('project_name', 'Proyecto')
    
    if not project_id:
        flash("No se proporcionó el ID del proyecto", "error")
        return redirect(url_for('history'))

    registros = []
    conn = None
    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()

        # 1. Obtener los registros principales (sin cambios aquí)
        cursor.execute("""
            SELECT id_registro, zona_intervencion, items, metros_lineales, proximas_tareas
            FROM registrosbitacoraeqing
            WHERE id_proyecto = %s
            ORDER BY id_registro DESC
        """, (project_id,))
        
        registros_principales = cursor.fetchall()

        # 2. Para cada registro, obtener sus fotos y videos CON SUS DESCRIPCIONES
        for row in registros_principales:
            id_registro = row[0]
            
            # --- LÓGICA DE FOTOS ACTUALIZADA ---
            # Ahora seleccionamos también la columna 'description'
            cursor.execute("SELECT imagen_base64, description FROM fotos_registro WHERE id_registro = %s", (id_registro,))
            fotos = []
            for item in cursor.fetchall():
                fotos.append({
                    'file_data': item[0],   # El archivo en base64
                    'description': item[1]  # La nueva descripción
                })
            
            # --- LÓGICA DE VIDEOS ACTUALIZADA ---
            # Hacemos lo mismo para los videos
            cursor.execute("SELECT video_base64, description FROM videos_registro WHERE id_registro = %s", (id_registro,))
            videos = []
            for item in cursor.fetchall():
                videos.append({
                    'file_data': item[0],   # El video en base64
                    'description': item[1]  # La nueva descripción
                })

            registros.append({
                'id': id_registro,
                'zona_intervencion': row[1],
                'items_value': row[2],
                'metros_lineales': row[3],
                'proximas_tareas': row[4],
                'fotos': fotos,   # Ahora es una lista de objetos
                'videos': videos  # Ahora es una lista de objetos
            })

    except Exception as e:
        print(f"Error al obtener registros: {str(e)}")
        flash("Error al cargar el historial de registros.", "error")
    finally:
        if conn:
            conn.close()
    
    return render_template('historialRegistro.html',
                           registros=registros,
                           project_name=project_name,
                           project_id=project_id)

@app.route('/disciplinerecords')
def disciplinerecords():
    return render_template('disciplinerecords.html')

@app.route('/projectdetails')
def projectdetails():
    return render_template('projectdetails.html')

@app.route('/addproject', methods=['GET', 'POST'])
def add_project():
    if 'user_id' not in session:  # Asegúrate de tener el user_id en la sesión
        return redirect(url_for('principalscreen'))
    
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            project_data = {
                'name': request.form['project-name'],
                'start_date': request.form['start-date'],
                'end_date': request.form['end-date'],
                'director': request.form['director'],
                'location': request.form['location'],
                'coordinates': request.form['coordinates'],
                'user_id': session['user_id']  # ID del usuario actual
            }
            
            # Guardar en PostgreSQL
            project_id = create_project(
                project_data['user_id'],
                project_data['name'],
                project_data['start_date'],
                project_data['end_date'],
                project_data['director'],
                project_data['location'],
                project_data['coordinates']
            )
            
            if project_id:
                flash('Proyecto creado exitosamente', 'success')
                return redirect(url_for('registros'))
            else:
                flash('Error al crear el proyecto', 'error')
                
        except Exception as e:
            flash(f'Error al guardar el proyecto: {str(e)}', 'error')
    
    return render_template('addproject.html')


@app.route('/ask', methods=['POST'])
def ask_question_route():
    data = request.json
    question = data.get('question', '')
    if not question:
        return jsonify({'error': 'No question provided'}), 400

    success = synthesize_speech(question)
    if success:
        return jsonify({'response': ''}), 200
    else:
        return jsonify({'error': 'Error al sintetizar la pregunta.'}), 500

@app.route('/guardar-registro', methods=['POST'])
def guardar_registro():
    try:
        data = request.get_json()
        print("Datos recibidos:", data)  # 🐞 DEBUG
        respuestas = data.get('respuestas')
        fotos = data.get('fotos', [])
        videos = data.get('videos', [])
        project_id = data.get('project_id')

        if not respuestas or not project_id:
            return jsonify({"error": "Faltan datos requeridos."}), 400

        # Guardar en PostgreSQL
        insert_registro_bitacora(respuestas, int(project_id), fotos, videos)

        return jsonify({"mensaje": "Registro guardado exitosamente!!"}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/eliminar-proyecto', methods=['POST'])
def eliminar_proyecto():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    data = request.get_json()
    proyecto_id = data.get('id_proyecto')

    if not proyecto_id:
        return jsonify({'error': 'Falta el ID del proyecto'}), 400

    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()

        # Asegurarse de que el proyecto pertenece al usuario
        cursor.execute("""
            DELETE FROM proyectos
            WHERE id_proyecto = %s AND user_id = %s
        """, (proyecto_id, session['user_id']))
        conn.commit()

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route('/transcribe-audio', methods=['POST'])
def transcribe_audio():
    try:
        if 'audio' not in request.files:
            print("🔴 No se recibió archivo de audio.")
            return jsonify({"error": "No se envió el archivo de audio"}), 400

        file = request.files['audio']
        print(f"📥 Recibido archivo: {file.filename}")

        # Guardar el archivo temporalmente
        temp_input = tempfile.NamedTemporaryFile(delete=False, suffix=".webm")
        file.save(temp_input.name)
        print(f"💾 Guardado en: {temp_input.name}")

        temp_wav = tempfile.NamedTemporaryFile(delete=False, suffix=".wav")
        formato_detectado = None

        try:
            print("🔍 Intentando decodificar como webm...")
            audio = AudioSegment.from_file(temp_input.name, format="webm")
            print("✅ Decodificado como webm.")
            formato_detectado = "webm"
        except Exception as e_webm:
            print("⚠️ Falla al decodificar como webm:", str(e_webm))
            try:
                print("🔁 Intentando decodificar como mp4...")
                audio = AudioSegment.from_file(temp_input.name, format="mp4")
                print("✅ Decodificado como mp4.")
                formato_detectado = "mp4"
            except Exception as e_mp4:
                print("❌ Fallo total al decodificar audio.")
                traceback.print_exc()
                return jsonify({
                    "error": "No se pudo procesar el audio.",
                    "error_webm": str(e_webm),
                    "error_mp4": str(e_mp4)
                }), 500

        # Exportar a WAV
        audio.export(temp_wav.name, format="wav")
        print("🔄 Exportado a WAV:", temp_wav.name)

        # Transcribir con Azure
        speech_config = get_speech_config()
        audio_config = speechsdk.audio.AudioConfig(filename=temp_wav.name)
        recognizer = speechsdk.SpeechRecognizer(speech_config=speech_config, audio_config=audio_config)
        result = recognizer.recognize_once_async().get()

        if result.reason == speechsdk.ResultReason.RecognizedSpeech:
            print("✅ Texto reconocido:", result.text)
            return jsonify({
                "text": result.text,
                "formato_detectado": formato_detectado
            })
        else:
            print("⚠️ No se reconoció el audio:", result.reason)
            return jsonify({
                "error": "No se reconoció el audio.",
                "formato_detectado": formato_detectado
            }), 400

    except Exception as e:
        print("❌ Error general en transcribe_audio:", str(e))
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

#Exportar registros seleccionados a Excel
@app.route('/exportar-registros-excel', methods=['POST'])
def exportar_registros_excel():
    registro_ids = request.form.getlist('registro_ids')
    project_id = request.form.get('project_id')

    if not registro_ids and not project_id:
        return "No se seleccionaron registros ni proyecto", 400

    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()

        if not registro_ids:
            cursor.execute("""
                SELECT id_registro, zona_intervencion, items, metros_lineales, proximas_tareas, foto_base64
                FROM registrosbitacoraeqing
                WHERE id_proyecto = %s
                ORDER BY id_registro DESC
            """, (project_id,))
        else:
            format_ids = tuple(map(int, registro_ids))
            cursor.execute("""
                SELECT id_registro, zona_intervencion, items, metros_lineales, proximas_tareas, foto_base64
                FROM registrosbitacoraeqing
                WHERE id_registro IN %s
                ORDER BY id_registro DESC
            """, (format_ids,))

        rows = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Registros"

        # Encabezado
        ws.append(["ID", "Zona de Intervención", "Ítems", "Metros Lineales", "Próximas Tareas", "Foto"])

        row_index = 2  # Comienza después del encabezado

        for row in rows:
            id_registro, zona, items, metros, tareas, foto_base64 = row
            ws.append([id_registro, zona, items, metros, tareas, ""])  # celda para imagen

            if foto_base64:
                try:
                    header, base64_data = foto_base64.split(',', 1) if ',' in foto_base64 else ('', foto_base64)
                    image_data = base64.b64decode(base64_data)
                    img = Image.open(io.BytesIO(image_data))
                    img.thumbnail((120, 120))  # redimensiona para celda
                    image_io = io.BytesIO()
                    img.save(image_io, format='PNG')
                    image_io.seek(0)

                    img_excel = ExcelImage(image_io)
                    img_excel.anchor = f"F{row_index}"
                    ws.add_image(img_excel)

                    # Ajustar altura de fila
                    ws.row_dimensions[row_index].height = 90
                except Exception as img_err:
                    print(f"Error al procesar imagen para registro {id_registro}: {img_err}")

            row_index += 1

        # Ajuste de anchos de columnas
        ws.column_dimensions['A'].width = 12  # ID
        ws.column_dimensions['B'].width = 30  # Zona de intervención
        ws.column_dimensions['C'].width = 25  # Ítems
        ws.column_dimensions['D'].width = 20  # Metros lineales
        ws.column_dimensions['E'].width = 35  # Próximas tareas
        ws.column_dimensions['F'].width = 18  # Imagen

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output,
                         download_name="registros_bitacora.xlsx",
                         as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        print(f"Error al exportar: {e}")
        return "Error al exportar", 500
    finally:
        if conn:
            conn.close()


@app.route('/exportar-proyectos-excel', methods=['POST'])
def exportar_proyectos_excel():
    project_ids = request.form.getlist('project_ids')
    
    if not project_ids:
        return "No se seleccionaron proyectos", 400

    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()
        wb = Workbook()
        wb.remove(wb.active)  # Eliminar hoja por defecto

        for pid in project_ids:
            try:
                pid_int = int(pid)
            except:
                continue

            # Obtener info del proyecto
            cursor.execute("""
                SELECT nombre_proyecto, fecha_inicio, fecha_fin, director_obra, ubicacion, coordenadas
                FROM proyectos WHERE id_proyecto = %s
            """, (pid_int,))
            proyecto = cursor.fetchone()
            if not proyecto:
                continue

            nombre, fecha_inicio, fecha_fin, director, ubicacion, coordenadas = proyecto
            sheet_title = (nombre[:30] or f"Proyecto {pid_int}").strip()
            ws = wb.create_sheet(title=sheet_title)

            # Encabezado de proyecto
            ws.append(["Nombre del Proyecto:", nombre])
            ws.append(["Fecha de Inicio:", str(fecha_inicio)])
            ws.append(["Fecha de Finalización:", str(fecha_fin)])
            ws.append(["Director del Proyecto:", director])
            ws.append(["Ubicación:", ubicacion])
            ws.append(["Coordenadas:", coordenadas])
            ws.append([])

            # Encabezado de registros
            ws.append(["ID", "Zona de Intervención", "Ítems Instalados", "Metros Lineales", "Próximas Tareas", "Foto"])

            # Obtener registros
            cursor.execute("""
                SELECT id_registro, zona_intervencion, items, metros_lineales, proximas_tareas, foto_base64
                FROM registrosbitacoraeqing
                WHERE id_proyecto = %s
                ORDER BY id_registro DESC
            """, (pid_int,))
            registros = cursor.fetchall()

            row_index = 9
            for registro in registros:
                idr, zona, items, metros, tareas, foto = registro
                ws.append([idr, zona, items, metros, tareas, ""])

                if foto:
                    try:
                        header, base64_data = foto.split(',', 1) if ',' in foto else ('', foto)
                        img_data = base64.b64decode(base64_data)
                        img = Image.open(io.BytesIO(img_data))
                        img.thumbnail((120, 120))
                        img_io = io.BytesIO()
                        img.save(img_io, format='PNG')
                        img_io.seek(0)

                        img_excel = ExcelImage(img_io)
                        img_excel.anchor = f"F{row_index}"
                        ws.add_image(img_excel)

                        ws.row_dimensions[row_index].height = 90
                    except Exception as e:
                        print(f"Error en imagen de registro {idr}: {e}")
                row_index += 1

            # Ajustes de columnas
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 35
            ws.column_dimensions['F'].width = 18

        # Generar archivo
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output,
                         download_name="proyectos_exportados.xlsx",
                         as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f"Error exportando proyectos: {e}")
        return "Error interno al exportar", 500
    finally:
        if conn:
            conn.close()

if __name__ == '__main__':
    app.run(debug=True)